import openpyxl as xl; 
import os;
  


files = [i for i in os.listdir("D:\\downloads\\ExcelFiles") if i.endswith("csv")]

for file in files: 
    # opening the source excel file 
    wb1 = xl.load_workbook(file) 
    ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
wb2 = xl.load_workbook('Sorted_Stocks.xlsx') 
ws2 = wb2.active 
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('Sorted_Stocks.xlsx')) 