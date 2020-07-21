from openpyxl import Workbook
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


wb2 = load_workbook('Sorted_Stocks.xlsx') 
ws2 = wb2.active 


for filename in os.listdir (os.path.expanduser('D:\\downloads\\ExcelFiles')):
    if filename.endswith(".csv"): 
        wb1 = load_workbook(filename) 
        ws1 = wb1.worksheets[0] 
        mr = ws1.max_row 
        mc = ws1.max_column
        for i in range (1, mr + 1): 
            for j in range (1, mc + 1): 
                c = ws1.cell(row = i, column = j) 
                ws2.cell(row = i, column = j).value = c.value
        continue
    else:
        continue


wb2.save('Sorted_Stocks.xlsx')